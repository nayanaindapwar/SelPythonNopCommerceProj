import openpyxl

class LoginData:

        test_LoginPage_data = [{"username": "admin@yourstore.com", "password": "admin"}, {"username": "abc@yourstore.com", "password": "abc"}]

        @staticmethod
        def getTestData(): # self parameter not required for static method
                data_list = []
                book = openpyxl.load_workbook(
                        "/home/nayana/PycharmProjects/DemoNopCommerce/TestData/LoginDataNopCommerce.xlsx")
                sheet = book.active
                for i in range(2, sheet.max_row + 1):
                        data_dict = {}
                        for j in range(2, sheet.max_column + 1):
                                data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

                        data_list.append(data_dict)

                return data_list
