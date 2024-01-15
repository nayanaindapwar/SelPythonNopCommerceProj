import openpyxl


class AddCustomerData:

        @staticmethod
        def getCustomerData(): # self parameter not required for static method
                add_data_list = []
                book = openpyxl.load_workbook(
                        "/home/nayana/PycharmProjects/DemoNopCommerce/TestData/AddCustomerDataNopCom.xlsx")
                sheet = book.active
                for i in range(2, sheet.max_row + 1):
                        add_data_dict = {}
                        for j in range(2, sheet.max_column + 1):
                                add_data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

                        add_data_list.append(add_data_dict)

                return add_data_list