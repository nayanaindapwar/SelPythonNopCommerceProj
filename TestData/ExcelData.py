from _ast import Dict

import openpyxl
book = openpyxl.load_workbook("/home/nayana/PycharmProjects/DemoNopCommerce/TestData/EditCustomerDataNopCom.xlsx")

sheet = book.active
Dict = {}
for i in range(2, sheet.max_row+1):
        for j in range(2, sheet.max_column+1):
            # print(sheet.cell(row=i, column=j).value)
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        print(Dict)

